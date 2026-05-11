import io
import os
from urllib.parse import unquote, urlparse
from django.conf import settings
from django.contrib.staticfiles import finders
from django.template.loader import get_template
from xhtml2pdf import pisa
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


def link_callback(uri, rel):
    """
    Convierte URLs de media/static en rutas locales para que xhtml2pdf pueda
    incrustar imágenes y recursos dentro del PDF.
    """
    parsed_uri = urlparse(uri)
    path = unquote(parsed_uri.path)

    if path.startswith(settings.MEDIA_URL):
        absolute_path = os.path.join(settings.MEDIA_ROOT, path.replace(settings.MEDIA_URL, "", 1))
    elif path.startswith(settings.STATIC_URL):
        relative_path = path.replace(settings.STATIC_URL, "", 1)
        static_root = getattr(settings, "STATIC_ROOT", "")
        absolute_path = finders.find(relative_path) or os.path.join(static_root, relative_path)
    else:
        absolute_path = uri

    if not os.path.isfile(absolute_path):
        return uri
    return absolute_path


def generate_pdf(template_path, context):
    """
    Genera un PDF a partir de un template HTML y un contexto.
    Retorna un objeto BytesIO con el contenido del PDF.
    """
    template = get_template(template_path)
    html = template.render(context)
    result = io.BytesIO()
    pdf = pisa.pisaDocument(io.BytesIO(html.encode("UTF-8")), result, link_callback=link_callback)
    if not pdf.err:
        result.seek(0)
        return result
    return None

def generate_excel_inventario(equipos_queryset):
    """
    Genera un archivo Excel (.xlsx) con el listado de equipos.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario de Equipos"

    # Encabezados
    headers = [
        "Código", "Nombre", "Tipo", "Marca",
        "Modelo", "Serie", "Área", "Estado técnico",
        "Disponibilidad", "Activo", "Fecha registro", "Última actualización"
    ]
    
    # Estilos para encabezados
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        
        # Ajustar ancho de columna (básico)
        ws.column_dimensions[cell.column_letter].width = 20

    # Datos
    for row_num, equipo in enumerate(equipos_queryset, 2):
        ws.cell(row=row_num, column=1, value=equipo.codigo_equipo)
        ws.cell(row=row_num, column=2, value=equipo.nombre_equipo)
        ws.cell(row=row_num, column=3, value=equipo.tipo_equipo.nombre if equipo.tipo_equipo else "N/A")
        ws.cell(row=row_num, column=4, value=equipo.marca.nombre if equipo.marca else "N/A")
        ws.cell(row=row_num, column=5, value=equipo.modelo)
        ws.cell(row=row_num, column=6, value=equipo.numero_serie or "S/N")
        ws.cell(row=row_num, column=7, value=str(equipo.area) if equipo.area else "Sin área")
        estado = equipo.estado_tecnico or equipo.estado
        ws.cell(row=row_num, column=8, value=estado.nombre if estado else "Sin estado")
        ws.cell(row=row_num, column=9, value=equipo.get_disponibilidad_display())
        ws.cell(row=row_num, column=10, value="Sí" if equipo.activo else "No")
        ws.cell(row=row_num, column=11, value=equipo.fecha_register.strftime("%d/%m/%Y %H:%M") if equipo.fecha_register else "")
        ws.cell(row=row_num, column=12, value=equipo.actualizado_en.strftime("%d/%m/%Y %H:%M") if equipo.actualizado_en else "")

    result = io.BytesIO()
    wb.save(result)
    result.seek(0)
    return result
